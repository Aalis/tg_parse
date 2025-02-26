from fastapi import FastAPI, HTTPException, Depends, Query, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm, OAuth2PasswordBearer
from fastapi.responses import RedirectResponse, StreamingResponse
from pydantic import BaseModel, EmailStr
from typing import List, Optional
import os
from dotenv import load_dotenv
from telethon import TelegramClient, errors
from telethon.tl.types import ChannelParticipantsAdmins, ChannelParticipantsRecent
from telethon.tl.functions.channels import GetFullChannelRequest
import asyncio
import re
import logging
from datetime import timedelta, datetime
from sqlalchemy.orm import Session
from zoneinfo import ZoneInfo
import io
from openpyxl import Workbook
import unicodedata
import urllib.parse

from app.core.token_pool import TokenPool
from app.database.session import get_db, engine, Base
from app.models.user import User
from app.models.telegram_data import TelegramGroup, TelegramMember
from app.auth.security import (
    get_password_hash,
    verify_password,
    create_access_token,
    get_current_active_user,
    ACCESS_TOKEN_EXPIRE_MINUTES,
)
from app.schemas.user import UserCreate, UserOut
from app.utils.verification import generate_verification_token, is_token_expired

# Create database tables
Base.metadata.create_all(bind=engine)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

app = FastAPI(title="Telegram Group Parser")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Models
class GroupInfo(BaseModel):
    group_id: str
    name: str
    member_count: Optional[int]
    description: Optional[str]

class MemberInfo(BaseModel):
    user_id: int
    username: Optional[str]
    first_name: Optional[str]
    last_name: Optional[str]
    is_premium: Optional[bool]
    can_message: bool
    is_admin: bool
    admin_title: Optional[str]

class GroupMembersResponse(BaseModel):
    success: bool
    data: Optional[List[MemberInfo]]
    error: Optional[str]
    total_count: Optional[int]
    has_more: bool

class TelegramResponse(BaseModel):
    success: bool
    data: Optional[dict]
    error: Optional[str]

class PoolStatus(BaseModel):
    total_tokens: int
    active_tokens: int
    tokens: List[dict]

# User registration and authentication models
class Token(BaseModel):
    access_token: str
    token_type: str

class UserResponse(BaseModel):
    id: int
    email: str
    username: str
    is_active: bool
    is_superuser: bool
    is_verified: bool

    class Config:
        orm_mode = True

class ExcelExportRequest(BaseModel):
    members: List[MemberInfo]
    group_name: str

# User registration and authentication endpoints
@app.post("/register", response_model=UserOut)
def register(user: UserCreate, db: Session = Depends(get_db)):
    # Check if user exists
    db_user = db.query(User).filter(
        (User.email == user.email) | (User.username == user.username)
    ).first()
    if db_user:
        if db_user.email == user.email:
            raise HTTPException(status_code=400, detail="Email already registered")
        else:
            raise HTTPException(status_code=400, detail="Username already taken")

    # Create new user
    hashed_password = get_password_hash(user.password)
    verification_token, token_expires = generate_verification_token()
    
    db_user = User(
        email=user.email,
        username=user.username,
        hashed_password=hashed_password,
        verification_token=verification_token,
        verification_token_expires=token_expires
    )
    
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    
    # In a real application, you would send an email here with the verification link
    # For this simple example, we'll just return the token in the response
    return {
        "id": db_user.id,
        "email": db_user.email,
        "username": db_user.username,
        "is_active": db_user.is_active,
        "is_verified": db_user.is_verified,
        "verification_token": verification_token  # In production, don't expose this
    }

@app.get("/verify/{token}")
def verify_email(token: str, db: Session = Depends(get_db)):
    """Verify a user's email address using the verification token."""
    user = db.query(User).filter(User.verification_token == token).first()
    if not user:
        raise HTTPException(status_code=404, detail="Invalid verification token")
    
    if user.is_verified:
        return RedirectResponse(url="http://localhost:5173/?verified=already")
    
    if is_token_expired(user.verification_token_expires):
        return RedirectResponse(url="http://localhost:5173/?verified=expired")
    
    user.is_verified = True
    user.verification_token = None
    user.verification_token_expires = None
    db.commit()
    
    return RedirectResponse(url="http://localhost:5173/?verified=success")

@app.post("/token", response_model=Token)
async def login_for_access_token(
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    # Find user
    user = db.query(User).filter(User.username == form_data.username).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Create access token
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/users/me", response_model=UserResponse)
async def read_users_me(current_user: User = Depends(get_current_active_user)):
    return current_user

# Initialize token pool
API_ID = os.getenv("TELEGRAM_API_ID")
API_HASH = os.getenv("TELEGRAM_API_HASH")

if not all([API_ID, API_HASH]):
    raise ValueError("TELEGRAM_API_ID and TELEGRAM_API_HASH must be set in environment variables")

token_pool = TokenPool(API_ID, API_HASH)

@app.on_event("startup")
async def startup_event():
    """Initialize all clients in the token pool when the FastAPI app starts"""
    try:
        await token_pool.initialize_clients()
        logger.info("Token pool initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize token pool: {e}")
        raise

@app.on_event("shutdown")
async def shutdown_event():
    """Close all client connections when the FastAPI app shuts down"""
    try:
        await token_pool.close_all()
        logger.info("All clients disconnected successfully")
    except Exception as e:
        logger.error(f"Error disconnecting clients: {e}")

def extract_group_identifier(group_link: str) -> str:
    """Extract group username or chat_id from various link formats."""
    group_link = group_link.strip()
    
    if "t.me/" in group_link:
        group_link = re.sub(r'^https?://', '', group_link)
        match = re.search(r"t\.me/([^/?]+)", group_link)
        if match:
            return match.group(1)
    
    if group_link.startswith("@"):
        return group_link[1:]
    
    if "/" not in group_link and "." not in group_link:
        return group_link
    
    return group_link

@app.get("/api/pool-status", response_model=PoolStatus)
async def get_pool_status():
    """Get the current status of the token pool"""
    return token_pool.get_pool_status()

@app.post("/api/parse-group", response_model=TelegramResponse)
async def parse_group(
    group_link: str = Query(..., description="Telegram group link or username"),
    db: Session = Depends(get_db)
):
    try:
        group_id = extract_group_identifier(group_link)
        logger.info(f"Attempting to parse group with identifier: {group_id}")
        
        if not group_id:
            return TelegramResponse(
                success=False,
                data=None,
                error="Invalid group link format"
            )
        
        client = await token_pool.get_client()
        if not client:
            return TelegramResponse(
                success=False,
                data=None,
                error="No available bot tokens"
            )
        
        try:
            # Get group info
            entity = await client.get_entity(group_id)
            logger.info(f"Found entity type: {type(entity).__name__}")
            
            # Get member count
            full_chat = await client(GetFullChannelRequest(entity))
            member_count = full_chat.full_chat.participants_count
            
            # Store the ID with -100 prefix for supergroups/channels
            proper_id = f"-100{entity.id}" if str(entity.id).isdigit() else str(entity.id)
            
            # Create group info without saving to database
            group_info = GroupInfo(
                group_id=proper_id,
                name=entity.title,
                member_count=member_count,
                description=getattr(entity, 'about', None)
            )
            
            return TelegramResponse(
                success=True,
                data=group_info.dict(),
                error=None
            )
            
        except errors.ChatAdminRequiredError:
            token_pool.mark_error(client)
            return TelegramResponse(
                success=False,
                data=None,
                error="Bot needs to be an admin of the group to access this information"
            )
        except errors.ChannelPrivateError:
            token_pool.mark_error(client)
            return TelegramResponse(
                success=False,
                data=None,
                error="This is a private group. The bot needs to be a member."
            )
        except ValueError as e:
            if "Could not find the input entity" in str(e):
                return TelegramResponse(
                    success=False,
                    data=None,
                    error="Group not found. Please check if the group exists and is accessible."
                )
            raise
            
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return TelegramResponse(
            success=False,
            data=None,
            error=f"An unexpected error occurred: {str(e)}"
        )

@app.get("/api/group-members/{group_id}", response_model=GroupMembersResponse)
async def get_group_members(
    group_id: str,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    db: Session = Depends(get_db)
):
    try:
        logger.info(f"Fetching members for group: {group_id} (offset: {offset}, limit: {limit})")
        
        client = await token_pool.get_client()
        if not client:
            return GroupMembersResponse(
                success=False,
                data=None,
                error="No available bot tokens",
                total_count=None,
                has_more=False
            )
        
        try:
            # Get the entity using proper channel ID handling
            try:
                if group_id.startswith('-100'):
                    from telethon.tl.types import PeerChannel
                    channel_id = int(group_id[4:])  # Remove the -100 prefix
                    # Create a PeerChannel object first
                    peer = PeerChannel(channel_id)
                    # Then get the full entity
                    entity = await client.get_entity(peer)
                else:
                    entity = await client.get_entity(group_id)
                
                logger.info(f"Successfully got entity of type: {type(entity).__name__}")
            except ValueError as e:
                logger.error(f"Error getting entity: {str(e)}")
                return GroupMembersResponse(
                    success=False,
                    data=None,
                    error="Could not find the group. Please check if the group exists and is accessible.",
                    total_count=None,
                    has_more=False
                )
            
            members = []
            total_count = 0
            
            # Get total member count
            try:
                full_chat = await client(GetFullChannelRequest(entity))
                total_count = full_chat.full_chat.participants_count
                logger.info(f"Total member count: {total_count}")
            except Exception as e:
                logger.warning(f"Could not get total member count: {str(e)}")
            
            # Get administrators first (always include them)
            admin_ids = set()
            if offset == 0:  # Only fetch admins for the first page
                try:
                    from telethon.tl.functions.channels import GetParticipantsRequest
                    from telethon.tl.types import ChannelParticipantsAdmins
                    
                    admins_result = await client(GetParticipantsRequest(
                        channel=entity,
                        filter=ChannelParticipantsAdmins(),
                        offset=0,
                        limit=100,  # Get all admins
                        hash=0
                    ))
                    
                    for admin in admins_result.users:
                        try:
                            member_info = MemberInfo(
                                user_id=admin.id,
                                username=getattr(admin, 'username', None),
                                first_name=getattr(admin, 'first_name', None),
                                last_name=getattr(admin, 'last_name', None),
                                is_premium=getattr(admin, 'premium', None),
                                can_message=bool(getattr(admin, 'username', None)),
                                is_admin=True,
                                admin_title=None
                            )
                            members.append(member_info)
                            admin_ids.add(admin.id)
                            
                        except Exception as e:
                            logger.warning(f"Error processing admin {admin.id}: {str(e)}")
                            continue
                except Exception as e:
                    logger.warning(f"Error fetching admins: {str(e)}")
            
            # Calculate how many regular members to fetch
            remaining_limit = limit
            if offset == 0:
                remaining_limit = max(0, limit - len(members))
                regular_offset = 0
            else:
                regular_offset = offset
            
            # Get regular members
            if remaining_limit > 0:
                try:
                    from telethon.tl.functions.channels import GetParticipantsRequest
                    from telethon.tl.types import ChannelParticipantsRecent
                    
                    participants_result = await client(GetParticipantsRequest(
                        channel=entity,
                        filter=ChannelParticipantsRecent(),
                        offset=regular_offset,
                        limit=remaining_limit,
                        hash=0
                    ))
                    
                    for participant in participants_result.users:
                        if participant.id not in admin_ids:
                            try:
                                member_info = MemberInfo(
                                    user_id=participant.id,
                                    username=getattr(participant, 'username', None),
                                    first_name=getattr(participant, 'first_name', None),
                                    last_name=getattr(participant, 'last_name', None),
                                    is_premium=getattr(participant, 'premium', None),
                                    can_message=bool(getattr(participant, 'username', None)),
                                    is_admin=False,
                                    admin_title=None
                                )
                                members.append(member_info)
                                
                            except Exception as e:
                                logger.warning(f"Error processing member {participant.id}: {str(e)}")
                                continue
                    
                except Exception as e:
                    logger.error(f"Error fetching regular members: {str(e)}")
                    return GroupMembersResponse(
                        success=False,
                        data=None,
                        error=f"Failed to fetch members: {str(e)}",
                        total_count=None,
                        has_more=False
                    )
            
            # Calculate if there are more members to fetch
            has_more = (offset + len(members)) < total_count if total_count > 0 else len(members) >= limit
            
            return GroupMembersResponse(
                success=True,
                data=members,
                error=None,
                total_count=total_count,
                has_more=has_more
            )
            
        except errors.ChatAdminRequiredError:
            token_pool.mark_error(client)
            return GroupMembersResponse(
                success=False,
                data=None,
                error="Bot needs to be an admin of the group to access member list",
                total_count=None,
                has_more=False
            )
        except errors.ChannelPrivateError:
            token_pool.mark_error(client)
            return GroupMembersResponse(
                success=False,
                data=None,
                error="This is a private group. The bot needs to be a member.",
                total_count=None,
                has_more=False
            )
            
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        return GroupMembersResponse(
            success=False,
            data=None,
            error=f"An unexpected error occurred: {str(e)}",
            total_count=None,
            has_more=False
        )

# Add endpoint to export group members as CSV
@app.get("/api/export-members/{group_id}")
async def export_members_csv(
    group_id: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    import csv
    from fastapi.responses import StreamingResponse
    from io import StringIO
    
    # Check if group exists and user has access
    db_group = db.query(TelegramGroup).filter(
        TelegramGroup.group_id == group_id,
        TelegramGroup.created_by_user_id == current_user.id
    ).first()
    
    if not db_group:
        raise HTTPException(
            status_code=404,
            detail="Group not found or access denied"
        )
    
    # Get all members for this group
    members = db.query(TelegramMember).filter(TelegramMember.group_id == db_group.id).all()
    
    # Create CSV file
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        "User ID", "Username", "First Name", "Last Name",
        "Is Premium", "Can Message", "Is Admin", "Admin Title"
    ])
    
    # Write member data
    for member in members:
        writer.writerow([
            member.user_id,
            member.username or "",
            member.first_name or "",
            member.last_name or "",
            member.is_premium,
            member.can_message,
            member.is_admin,
            member.admin_title or ""
        ])
    
    # Prepare the response
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=group_{group_id}_members.csv"
        }
    )

def sanitize_filename(filename: str) -> str:
    """Sanitize filename to remove special characters and emojis while preserving readability."""
    # Normalize unicode characters
    filename = unicodedata.normalize('NFKD', filename)
    
    # Remove emojis and other special characters but keep basic punctuation
    filename = ''.join(c for c in filename if not unicodedata.combining(c) 
                      and (c.isalnum() or c in ' -_()[]{}.,'))
    
    # Replace multiple spaces/hyphens with single underscore
    filename = re.sub(r'[-\s]+', '_', filename)
    
    # Remove any non-allowed characters that might remain
    filename = re.sub(r'[^\w\s-]', '', filename)
    
    # Trim underscores from start and end
    filename = filename.strip('_')
    
    # Ensure the filename is not empty
    if not filename:
        filename = "group"
        
    return filename

@app.post("/api/export-excel")
async def export_excel(
    data: ExcelExportRequest,
    current_user: User = Depends(get_current_active_user)
):
    """Export members data to Excel file."""
    if not current_user.is_verified:
        raise HTTPException(
            status_code=403,
            detail="Account not verified. Please verify your email first."
        )

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"

    # Write headers
    headers = ['User ID', 'Username', 'First Name', 'Last Name', 'Is Premium', 'Is Admin', 'Admin Title', 'Can Message']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row, member in enumerate(data.members, 2):
        ws.cell(row=row, column=1, value=member.user_id)
        ws.cell(row=row, column=2, value=member.username or '')
        ws.cell(row=row, column=3, value=member.first_name or '')
        ws.cell(row=row, column=4, value=member.last_name or '')
        ws.cell(row=row, column=5, value='Yes' if member.is_premium else 'No')
        ws.cell(row=row, column=6, value='Yes' if member.is_admin else 'No')
        ws.cell(row=row, column=7, value=member.admin_title or '')
        ws.cell(row=row, column=8, value='Yes' if member.can_message else 'No')

    # Style the header row
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
        
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create filename with group name and current date
    current_date = datetime.now().strftime("%Y%m%d")
    safe_filename = sanitize_filename(data.group_name)
    filename = f"{safe_filename}-members-{current_date}.xlsx"
    encoded_filename = urllib.parse.quote(filename)
    
    # Return the Excel file as a downloadable response with properly encoded filename
    headers = {
        'Content-Disposition': f'attachment; filename*=UTF-8\'\'{encoded_filename}'
    }
    
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True) 